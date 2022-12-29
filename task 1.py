import csv
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, numbers


class DictByCity:
    def __init__(self):
        self.dic_data = {}
        self.total_cnt = 0

    def addata(self, city, val):
        if city not in self.dic_data.keys():
            self.dic_data[city] = [val, 1]
        else:
            self.dic_data[city][0] += val
            self.dic_data[city][1] += 1
        self.total_cnt += 1

    def get_sal(self):
        sub_dic = {}
        dic_sorted = dict(sorted(self.dic_data.items(), key=lambda item: item[1][0] // item[1][1], reverse=True))
        for key, val in dic_sorted.items():
            ratio = val[1] / self.total_cnt
            if ratio >= 0.01:
                sub_dic[key] = int(val[0] // val[1])
            if len(sub_dic) == 10:
                break
        return sub_dic

    def get_cnt(self):
        sub_dic = {}
        dic_sorted = dict(sorted(self.dic_data.items(), key=lambda item: item[1][1], reverse=True))
        for key, val in dic_sorted.items():
            ratio = val[1] / self.total_cnt
            if ratio >= 0.01:
                sub_dic[key] = round(ratio, 4)
            if len(sub_dic) == 10:
                break
        return sub_dic


class DictByYear:
    def __init__(self):
        self.dic_data = {}

    def addata(self, year, val, cnt):
        if year not in self.dic_data.keys():
            self.dic_data[year] = [val, cnt]
        else:
            self.dic_data[year][0] += val
            self.dic_data[year][1] += cnt

    def get_sal(self):
        sub_dic = {}
        for key, val in self.dic_data.items():
            sub_dic[key] = int(val[0] // val[1]) if val[1] != 0 else 0
        return sub_dic

    def get_cnt(self):
        sub_dic = {}
        for key, val in self.dic_data.items():
            sub_dic[key] = val[1]
        return sub_dic


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vac_obj = DataSet.set_vac(*DataSet.csv_reader(file_name))

        self.sal_year = DictByYear()
        self.sal_job_year = DictByYear()
        self.sal_cnt_city = DictByCity()

    @staticmethod
    def csv_reader(file_name):
        file_csv = csv.reader(open(file_name, encoding='utf_8_sig'))
        data_list = [x for x in file_csv if x.count('') == 0]
        return data_list[0], data_list[1:]

    @staticmethod
    def set_vac(headers, vacancies):
        vac_list = []
        for vac in vacancies:
            dict = {}
            for i, items in enumerate(headers):
                dict[items] = vac[i]
            vac_list.append(
                Vacancy(dict['name'], dict['salary_from'], dict['salary_to'], dict['salary_currency'], dict['area_name'],
                        dict['published_at']))
        return vac_list

    def collect_stat(self, proff):
        for vac in self.vac_obj:
            self.sal_year.addata(vac.published_at.year, vac.average_sal, 1)
            if proff in vac.name:
                self.sal_job_year.addata(vac.published_at.year, vac.average_sal, 1)
            else:
                self.sal_job_year.addata(vac.published_at.year, 0, 0)
            self.sal_cnt_city.addata(vac.area_name, vac.average_sal)

    def get_stat(self):
        return self.sal_year.get_sal(), \
               self.sal_year.get_cnt(), \
               self.sal_job_year.get_sal(), \
               self.sal_job_year.get_cnt(), \
               self.sal_cnt_city.get_sal(), \
               self.sal_cnt_city.get_cnt()

    def print_stat(self):
        print(f"Динамика уровня зарплат по годам: {self.sal_year.get_sal()}")
        print(f"Динамика количества вакансий по годам: {self.sal_year.get_cnt()}")
        print(
            f"Динамика уровня зарплат по годам для выбранной профессии: {self.sal_job_year.get_sal()}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {self.sal_job_year.get_cnt()}")
        print(f"Уровень зарплат по городам (в порядке убывания): {self.sal_cnt_city.get_sal()}")
        print(f"Доля вакансий по городам (в порядке убывания): {self.sal_cnt_city.get_cnt()}")


class Vacancy:
    cur_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.average_sal = (float(salary_from) + float(salary_to)) / 2 * Vacancy.cur_rub[
            salary_currency]
        self.area_name = area_name
        self.published_at = datetime.datetime.strptime(published_at, "%Y-%m-%dT%H:%M:%S%z")


class Report:
    def __init__(self, sal_year, cnt_year, sal_job_year, job_cnt_yaer, sal_city, cnt_city, prof):
        self.prof = prof
        self.sal_year = sal_year
        self.cnt_year = cnt_year
        self.prof_sal_year = sal_job_year
        self.prof_cnt_year = job_cnt_yaer
        self.sal_city = sal_city
        self.cnt_city = cnt_city

    @staticmethod
    def as_text(val):
        if val is None:
            return ""
        return str(val)

    def generate_excel(self):
        wb1 = Workbook()
        thin = Side(border_style="thin", color="000000")
        self.create_tab1(wb1, thin)
        self.create_tab2(wb1, thin)
        wb1.save("report.xlsx")

    def create_tab1(self, workbook, pen):
        tab1 = workbook.active
        tab1.title = "Статистика по годам"
        heads = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.prof}", "Количество вакансий",
                 f"Количество вакансий - {self.prof}"]
        for i, head in enumerate(heads):
            tab1.cell(row=1, column=(1 + i), value=head).font = Font(bold=True)
        for year, value in self.sal_year.items():
            tab1.append([year, value, self.prof_sal_year[year], self.cnt_year[year],
                         self.prof_cnt_year[year]])
        Report.set_len_and_border(tab1, pen)

    def create_tab2(self, workbook, pen):
        tab2 = workbook.create_sheet("Статистика по городам")
        heads = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        for i, head in enumerate(heads):
            tab2.cell(row=1, column=1 + i, value=head).font = Font(bold=True)
        for i, (city, value) in enumerate(self.sal_city.items()):
            city_count, value_count = list(self.cnt_city.items())[i]
            tab2.append([city, value, "", city_count, value_count])
            tab2.cell(row=2+i, column=5).number_format = numbers.BUILTIN_FORMATS[10]
        Report.set_len_and_border(tab2, pen, True)

    @staticmethod
    def set_len_and_border(tab, pen, is_city=False):
        for i, column in enumerate(tab.columns):
            len1 = max(len(Report.as_text(cell.value)) for cell in column)
            tab.column_dimensions[column[0].column_letter].width = len1 + 2
            for cell in column:
                if is_city and i == 2:
                    break
                cell.border = Border(left=pen, right=pen, top=pen, bottom=pen)

file_name = input('Введите название файла: ')
proff = input('Введите название профессии: ')
data1 = DataSet(file_name)
data1.collect_stat(proff)
data1.print_stat()

rep = Report(*data1.get_stat(), proff)
rep.generate_excel()